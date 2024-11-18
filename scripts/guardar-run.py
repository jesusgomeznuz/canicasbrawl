import pandas as pd
from openpyxl import load_workbook
import os
import shutil
from glob import glob

def guardar_run():
    # Step 1: Create a new folder in "D:\canicasbrawl\Runs"
    runs_path = r"D:\canicasbrawl\Runs"
    existing_folders = [int(folder) for folder in os.listdir(runs_path) if folder.isdigit()]
    new_folder_number = max(existing_folders, default=0) + 1
    new_folder_path = os.path.join(runs_path, str(new_folder_number))
    os.makedirs(new_folder_path)    
    print(f"Folder created: {new_folder_path}")

    # Step 2: Find the most recent run video in "C:\Users\LENOVO\HowToMakeAVideoGame\HowToMakeAVideoGame\Recordings"
    recordings_path = r"C:\Users\LENOVO\HowToMakeAVideoGame\HowToMakeAVideoGame\Recordings"
    video_files = glob(os.path.join(recordings_path, "Movie_*.mp4"))
    latest_video = max(video_files, key=os.path.getctime)
    print(f"Most recent video found: {latest_video}")

    # Step 3: Copy and paste the video into the new subfolder
    shutil.copy(latest_video, new_folder_path)
    print(f"Video copied to: {new_folder_path}")

    # Step 4: Delete the winner_log.csv file in "D:\canicasbrawl\scripts" if it exists
    scripts_path = r"D:\canicasbrawl\scripts"
    winner_log_script_path = os.path.join(scripts_path, "winner_log.csv")
    if os.path.exists(winner_log_script_path):
        os.remove(winner_log_script_path)
        print(f"File deleted: {winner_log_script_path}")

    # Step 5: Copy the winner_log.csv file to the new subfolder and to "D:\canicasbrawl\scripts"
    winner_log_source_path = r"C:\Users\LENOVO\AppData\LocalLow\DefaultCompany\HowToMakeAVideoGame\winner_log.csv"
    shutil.copy(winner_log_source_path, new_folder_path)
    shutil.copy(winner_log_source_path, scripts_path)
    print(f"File winner_log.csv copied to: {new_folder_path} and {scripts_path}")
    
    # Copy the dual_winner_log.csv file to the new subfolder and to "scripts"
    dual_winner_log_source_path = r"C:\Users\LENOVO\AppData\LocalLow\DefaultCompany\HowToMakeAVideoGame\dual_winner_log.csv"
    if os.path.exists(dual_winner_log_source_path):
        shutil.copy(dual_winner_log_source_path, new_folder_path)
        shutil.copy(dual_winner_log_source_path, scripts_path)
        print(f"File dual_winner_log.csv copied to: {new_folder_path} and {scripts_path}")
    else:
        print(f"File dual_winner_log.csv not found.")

# Run the function to save the run
guardar_run()

# Read the CSV file
file_path = r"D:\canicasbrawl\scripts\winner_log.csv"
df = pd.read_csv(file_path)

# Remove duplicates
df = df.drop_duplicates()

# Convert time columns to milliseconds for easier calculations
def time_to_ms(time_str):
    h, m, s, ms = map(int, time_str.split(':'))
    return ((h * 60 + m) * 60 + s) * 1000 + ms

df['Time_ms'] = df['Time'].apply(time_to_ms)

# Determine the winner based on the highest last time record
max_time_row = df.loc[df['Time_ms'].idxmax()]
winner = max_time_row['Nickname']

# Save the winner.csv file
winner_file_path = r"D:\canicasbrawl\scripts\winner.csv"
with open(winner_file_path, 'w') as f:
    f.write('nickname\n')
    f.write(winner)

# Calculate the time each player was in the lead
next_time_ms = list(df['Time_ms'][1:]) + [df['Time_ms'].iloc[-1]]
df['NextTime_ms'] = next_time_ms
df['LeadTime_ms'] = df['NextTime_ms'] - df['Time_ms']

# Filter out invalid records
df = df.iloc[:-1]  # Remove the last record that doesn't have a next time

# Group by Nickname and sum the times
total_times = {}
for idx, row in df.iterrows():
    nickname = row['Nickname']
    lead_time = row['LeadTime_ms']
    if nickname in total_times:
        total_times[nickname] += lead_time
    else:
        total_times[nickname] = lead_time

# Convert to seconds and prepare the results
results = pd.DataFrame(list(total_times.items()), columns=['Nickname', 'TotalTime'])
results['TotalTime'] = results['TotalTime'] / 1000

# Verify that winner.csv exists and contains data
if not os.path.exists(winner_file_path) or os.path.getsize(winner_file_path) == 0:
    raise FileNotFoundError(f"The file {winner_file_path} does not exist or is empty.")

# Read winner.csv to get the winner's name
winner_df = pd.read_csv(winner_file_path, encoding='latin1')
winner = winner_df.loc[0, 'nickname']

# Add 10 seconds to the winner
results.loc[results['Nickname'] == winner, 'TotalTime'] += 10

# Save the results.csv file
results_file_path = r"D:\canicasbrawl\scripts\results.csv"
results.to_csv(results_file_path, index=False)

print("Files 'results.csv' and 'winner.csv' generated successfully.")

def procesar_resultados():
    try:
        # Read the preferences file to get the list and number of players
        preferencias_path = r"D:\canicasbrawl\scripts\preferencias.csv"
        preferencias_df = pd.read_csv(preferencias_path)
        num_jugadores = len(preferencias_df)
        last_player = preferencias_df['nickname'].iloc[-1].strip().lower()  # Last player in the list

        # Read the results.csv file
        results_path = r"D:\canicasbrawl\scripts\results.csv"
        results_df = pd.read_csv(results_path)

        # Convert names to lowercase in results_df
        results_df['Nickname'] = results_df['Nickname'].str.strip().str.lower()

        # Read the existing Excel file
        excel_path = r"D:\canicasbrawl\historico_runs.xlsx"
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # Determine the column for the new run
        run_col = sheet.max_column + 1
        run_col_letter = chr(ord('A') + run_col - 1)
        run_col_header = f"Run{run_col - 1}"
        sheet.cell(row=1, column=run_col, value=run_col_header)
        print(f"Processing results in column {run_col_letter} ({run_col_header}).")

        # Create a dictionary to quickly lookup times by nickname
        results_dict = dict(zip(results_df['Nickname'], results_df['TotalTime']))

        # Debug: Verify that the names and times in results.csv are correct
        print(f"Results obtained: {results_dict}")

        # Update players' times in the Excel sheet
        last_player_row = None
        for row in range(2, sheet.max_row + 1):
            nickname = sheet.cell(row=row, column=1).value.strip().lower()
            if nickname in results_dict:
                total_time = results_dict[nickname]
                sheet.cell(row=row, column=run_col, value=total_time)
                print(f"Writing {total_time} for {nickname} in row {row}.")
            else:
                sheet.cell(row=row, column=run_col, value=0)
                print(f"Writing 0 for {nickname} in row {row}.")

            # Update the row of the last processed player
            if nickname == last_player:
                last_player_row = row

        # Confirm that all players received a time
        if last_player_row is not None:
            print(f"Last player processed: {last_player} in row {last_player_row}.")
        else:
            print("Error: Last player not found in the Excel sheet.")

        # Read winner.csv to get the winner's name
        winner_path = r"D:\canicasbrawl\scripts\winner.csv"
        with open(winner_path, 'r', encoding='latin1') as f:
            winner = f.readlines()[1].strip().lower()
        print(f"Winner obtained: {winner.capitalize()}.")

        # Place the winner just below the last processed player
        if last_player_row is not None:
            winner_row = last_player_row + 1
            sheet.cell(row=winner_row, column=1, value='Winner')
            sheet.cell(row=winner_row, column=run_col, value=winner.capitalize())
            print(f"Writing {winner.capitalize()} as winner in row {winner_row}.")

        # Save the updated Excel file
        workbook.save(excel_path)
        print(f"Results successfully saved in {excel_path}.")

    except Exception as e:
        print(f"Error processing results: {e}")

# Run the function to process the results
procesar_resultados()

def mover_mas_archivos():
    # Define the path of the runs folder
    runs_path = r"D:\canicasbrawl\Runs"
    
    # Find the subfolder with the highest number
    existing_folders = [int(folder) for folder in os.listdir(runs_path) if folder.isdigit()]
    if not existing_folders:
        print("No folders in the runs directory.")
        return

    latest_folder_number = max(existing_folders)
    latest_folder_path = os.path.join(runs_path, str(latest_folder_number))
    print(f"Most recent folder found: {latest_folder_path}")

    # Define the path of the scripts folder
    scripts_path = r"D:\canicasbrawl\scripts"
    
    # Define the paths of the files to copy
    winner_csv_path = os.path.join(scripts_path, "winner.csv")
    results_csv_path = os.path.join(scripts_path, "results.csv")
    dual_winner_log_path = os.path.join(scripts_path, "dual_winner_log.csv")  # Add path for dual_winner_log.csv
    
    # Copy the files to the destination folder
    if os.path.exists(winner_csv_path):
        shutil.copy(winner_csv_path, latest_folder_path)
        print(f"File winner.csv copied to: {latest_folder_path}")
    if os.path.exists(results_csv_path):
        shutil.copy(results_csv_path, latest_folder_path)
        print(f"File results.csv copied to: {latest_folder_path}")
    if os.path.exists(dual_winner_log_path):
        shutil.copy(dual_winner_log_path, latest_folder_path)
        print(f"File dual_winner_log.csv copied to: {latest_folder_path}")

# Run the function to move the files
mover_mas_archivos()
